"""Trade Report Parser — handles XLSX, CSV, PDF from any broker.

Uses fuzzy column matching to identify: stock name/symbol, trade type,
quantity, price/value, execution date regardless of broker-specific headers.
"""

from __future__ import annotations

import io
import logging
import re
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

from sqlalchemy.ext.asyncio import AsyncSession

from backend.models.orm import TradeHistory

logger = logging.getLogger(__name__)

# Column name patterns for fuzzy matching (case-insensitive)
COLUMN_PATTERNS = {
    "ticker": [
        r"symbol", r"trading.?symbol", r"scrip", r"stock.?symbol", r"ticker",
        r"instrument", r"script", r"nse.?symbol",
    ],
    "stock_name": [
        r"stock.?name", r"company", r"name", r"scrip.?name", r"instrument.?name",
    ],
    "isin": [r"isin"],
    "trade_type": [
        r"type", r"buy.?sell", r"transaction.?type", r"trade.?type", r"order.?type",
        r"side", r"action",
    ],
    "quantity": [
        r"quantity", r"qty", r"shares", r"units", r"no.?of.?shares", r"traded.?qty",
    ],
    "price": [
        r"price", r"rate", r"avg.?price", r"trade.?price", r"execution.?price",
        r"average.?price",
    ],
    "value": [
        r"value", r"amount", r"trade.?value", r"total.?value", r"net.?amount",
    ],
    "exchange": [r"exchange", r"mkt", r"market", r"segment"],
    "order_id": [
        r"order.?id", r"exchange.?order", r"trade.?id", r"ref", r"reference",
    ],
    "executed_at": [
        r"date", r"time", r"execution.?date", r"trade.?date", r"order.?date",
        r"executed", r"timestamp", r"fill.?date",
    ],
    "status": [r"status", r"order.?status"],
}


def _match_column(header: str, field: str) -> bool:
    """Check if a header matches a field's patterns."""
    header_clean = re.sub(r"[^a-z0-9]", "", header.lower())
    for pattern in COLUMN_PATTERNS.get(field, []):
        pattern_clean = pattern.replace(".", "").replace("?", "").replace(r"\s", "")
        if re.search(pattern, header.lower()):
            return True
        # Also try without special chars
        if pattern_clean in header_clean:
            return True
    return False


def _identify_columns(headers: list[str]) -> dict[str, int]:
    """Map field names to column indices using fuzzy matching."""
    mapping: dict[str, int] = {}

    for field in COLUMN_PATTERNS:
        for i, header in enumerate(headers):
            if _match_column(header, field):
                if field not in mapping:  # First match wins
                    mapping[field] = i
                break

    return mapping


def _parse_date(value: Any) -> datetime | None:
    """Parse various date formats from broker reports."""
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value

    s = str(value).strip()
    if not s:
        return None

    # Common Indian broker date formats
    formats = [
        "%d-%m-%Y %I:%M %p",      # 17-11-2020 09:48 AM (Groww)
        "%d-%m-%Y %H:%M:%S",      # 17-11-2020 09:48:00
        "%d-%m-%Y %H:%M",         # 17-11-2020 09:48
        "%d-%m-%Y",               # 17-11-2020
        "%Y-%m-%d %H:%M:%S",      # 2020-11-17 09:48:00
        "%Y-%m-%d",               # 2020-11-17
        "%d/%m/%Y %H:%M:%S",      # 17/11/2020 09:48:00
        "%d/%m/%Y",               # 17/11/2020
        "%d %b %Y",               # 17 Nov 2020
        "%d-%b-%Y",               # 17-Nov-2020
        "%m/%d/%Y",               # 11/17/2020 (US format)
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt)
            return dt.replace(tzinfo=timezone.utc)
        except ValueError:
            continue

    return None


def _parse_trade_type(value: Any) -> str | None:
    """Normalize trade type to BUY or SELL."""
    if value is None:
        return None
    s = str(value).upper().strip()
    if s in ("BUY", "B", "BOUGHT", "PURCHASE"):
        return "BUY"
    if s in ("SELL", "S", "SOLD", "SALE"):
        return "SELL"
    return s if s else None


def _parse_number(value: Any) -> Decimal | None:
    """Parse a number from various formats."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    s = str(value).strip().replace(",", "").replace("₹", "").replace("Rs", "").replace(" ", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except Exception:
        return None


async def parse_xlsx(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse XLSX/XLS file into trade records."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    records = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (first row with multiple non-empty cells that match known patterns)
        header_idx = None
        headers = []
        for i, row in enumerate(rows):
            cells = [str(c).strip() if c else "" for c in row]
            non_empty = [c for c in cells if c]
            if len(non_empty) >= 4:
                # Check if it looks like a header
                match_count = sum(
                    1 for c in cells
                    if any(_match_column(c, f) for f in COLUMN_PATTERNS)
                )
                if match_count >= 3:
                    header_idx = i
                    headers = cells
                    break

        if header_idx is None:
            continue

        col_map = _identify_columns(headers)
        if "ticker" not in col_map and "stock_name" not in col_map:
            continue

        # Parse data rows
        for row in rows[header_idx + 1:]:
            cells = [c for c in row]
            if not cells or all(c is None for c in cells):
                continue

            record = _extract_record(cells, col_map, filename)
            if record:
                records.append(record)

    return records


async def parse_csv(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse CSV file into trade records."""
    import csv

    text = file_bytes.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if not rows:
        return []

    # Find header row
    header_idx = None
    headers = []
    for i, row in enumerate(rows):
        cells = [str(c).strip() for c in row]
        match_count = sum(
            1 for c in cells
            if any(_match_column(c, f) for f in COLUMN_PATTERNS)
        )
        if match_count >= 3:
            header_idx = i
            headers = cells
            break

    if header_idx is None:
        return []

    col_map = _identify_columns(headers)
    if "ticker" not in col_map and "stock_name" not in col_map:
        return []

    records = []
    for row in rows[header_idx + 1:]:
        if not row or all(not c.strip() for c in row):
            continue
        record = _extract_record(row, col_map, filename)
        if record:
            records.append(record)

    return records


def _extract_record(cells: list, col_map: dict[str, int], filename: str) -> dict | None:
    """Extract a single trade record from a row using the column mapping."""
    def get(field: str):
        idx = col_map.get(field)
        if idx is None or idx >= len(cells):
            return None
        return cells[idx]

    ticker = get("ticker")
    stock_name = get("stock_name")
    trade_type = _parse_trade_type(get("trade_type"))
    quantity = _parse_number(get("quantity"))
    price = _parse_number(get("price"))
    value = _parse_number(get("value"))
    executed_at = _parse_date(get("executed_at"))
    isin = str(get("isin") or "").strip() or None
    exchange = str(get("exchange") or "").strip() or None
    order_id = str(get("order_id") or "").strip() or None
    status = str(get("status") or "").strip().lower()

    # Skip non-executed orders
    if status and status not in ("executed", "filled", "complete", "success", ""):
        return None

    # Must have at minimum: ticker/name + type + quantity
    ticker_val = str(ticker or "").strip()
    if not ticker_val and stock_name:
        ticker_val = str(stock_name).strip()

    if not ticker_val or not trade_type or not quantity:
        return None

    # If price not available, derive from value/quantity
    if not price and value and quantity and quantity > 0:
        price = value / quantity

    return {
        "ticker": ticker_val.upper(),
        "isin": isin,
        "trade_type": trade_type,
        "quantity": float(quantity),
        "price": float(price) if price else 0,
        "value": float(value) if value else None,
        "exchange": exchange,
        "order_id": order_id,
        "executed_at": executed_at.isoformat() if executed_at else None,
        "source_file": filename,
    }


async def store_trades(
    db: AsyncSession, user_id: UUID, records: list[dict], broker: str | None = None
) -> int:
    """Store parsed trade records in DB. Returns count stored."""
    stored = 0
    for rec in records:
        executed_at = None
        if rec.get("executed_at"):
            try:
                executed_at = datetime.fromisoformat(rec["executed_at"])
            except (ValueError, TypeError):
                pass

        trade = TradeHistory(
            id=uuid4(),
            user_id=user_id,
            ticker=rec["ticker"],
            isin=rec.get("isin"),
            trade_type=rec["trade_type"],
            quantity=Decimal(str(rec["quantity"])),
            price=Decimal(str(rec["price"])),
            value=Decimal(str(rec["value"])) if rec.get("value") else None,
            exchange=rec.get("exchange"),
            order_id=rec.get("order_id"),
            executed_at=executed_at,
            broker=broker,
            source_file=rec.get("source_file"),
        )
        db.add(trade)
        stored += 1

    await db.commit()
    return stored


async def get_purchase_dates(db: AsyncSession, user_id: UUID, portfolio_id=None) -> dict[str, str]:
    """Get earliest BUY date per ticker from trade history.

    Returns {ticker: date_iso_string} for use in dividend calculations.
    """
    from sqlalchemy import select, func

    filters = [
        TradeHistory.user_id == user_id,
        TradeHistory.trade_type == "BUY",
        TradeHistory.executed_at.isnot(None),
    ]
    if portfolio_id:
        filters.append(TradeHistory.portfolio_id == portfolio_id)

    stmt = (
        select(
            TradeHistory.ticker,
            func.min(TradeHistory.executed_at).label("first_buy"),
        )
        .where(*filters)
        .group_by(TradeHistory.ticker)
    )
    result = await db.execute(stmt)
    return {
        row[0]: row[1].strftime("%Y-%m-%d") for row in result.all() if row[1]
    }
